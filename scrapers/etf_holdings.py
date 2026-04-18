"""
ETF holdings tracker.

Scrapes daily holdings for selected active ETFs and computes diff
against the previous trading day.

Tracked ETFs:
  00981A  主動統一台股增長      (ezmoney, fundCode=49YTW)
  00988A  主動統一全球創新      (ezmoney, fundCode=61YTW)
  00992A  群益台灣科技創新主動  (capitalfund, fundId=500)
"""

import json
import re
import html as html_mod
import os
import sys
from datetime import date

from db.connection import get_cursor

if getattr(sys, "frozen", False):
    # PyInstaller: anchor to the exe's directory so reports land next to daily_update.exe
    _BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(sys.executable)))
else:
    _BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
REPORT_DIR = os.path.join(_BASE_DIR, "reports")
from utils.http_client import fetch, get_session, _get_domain, _wait_for_rate_limit

# ---------------------------------------------------------------------------
# ETF registry
# ---------------------------------------------------------------------------

ETF_REGISTRY = [
    {
        "etf_id": "00981A",
        "source": "ezmoney",
        "fund_code": "49YTW",
    },
    {
        "etf_id": "00988A",
        "source": "ezmoney",
        "fund_code": "61YTW",
    },
    {
        "etf_id": "00991A",
        "source": "fhtrust",
        "etf_code": "ETF23",
    },
    {
        "etf_id": "00980A",
        "source": "nomura",
        "fund_no": "00980A",
    },
    {
        "etf_id": "00993A",
        "source": "allianz",
        "fund_id": "E0001",
    },
    {
        "etf_id": "00982A",
        "source": "capitalfund",
        "fund_id": "399",
    },
    {
        "etf_id": "00992A",
        "source": "capitalfund",
        "fund_id": "500",
    },
]

# ---------------------------------------------------------------------------
# Ezmoney (統一投信) — holdings embedded in HTML data-content
# ---------------------------------------------------------------------------

EZMONEY_URL = "https://www.ezmoney.com.tw/ETF/Fund/Info"


def _fetch_ezmoney(fund_code: str) -> list[dict]:
    """
    Fetch holdings from ezmoney HTML page.
    Returns list of {stock_id, stock_name, shares, weight}.
    """
    resp = fetch(EZMONEY_URL, params={"fundCode": fund_code})
    if resp is None:
        print(f"  [ERROR] Failed to fetch ezmoney page for {fund_code}")
        return []

    page = resp.text

    # Extract DataAsset JSON from data-content attribute
    m = re.search(r'id="DataAsset"\s+data-content="([^"]+)"', page)
    if not m:
        print(f"  [ERROR] DataAsset not found in page for {fund_code}")
        return []

    raw = html_mod.unescape(m.group(1))
    assets = json.loads(raw)

    # Find stock (ST) asset type
    holdings = []
    for asset in assets:
        if asset.get("AssetCode") != "ST":
            continue
        details = asset.get("Details") or []
        for d in details:
            stock_id = (d.get("DetailCode") or "").strip()
            stock_name = (d.get("DetailName") or "").strip()
            shares = d.get("Share")
            weight = d.get("NavRate")
            if stock_id and shares is not None:
                holdings.append({
                    "stock_id": stock_id,
                    "stock_name": stock_name,
                    "shares": int(shares),
                    "weight": float(weight) if weight is not None else None,
                })

    return holdings

# ---------------------------------------------------------------------------
# Capitalfund (群益投信) — JSON API
# ---------------------------------------------------------------------------

CAPITALFUND_URL = "https://www.capitalfund.com.tw/CFWeb/api/etf/buyback"


def _fetch_capitalfund(fund_id: str) -> list[dict]:
    """
    Fetch holdings from capitalfund API (POST with JSON body).
    Returns list of {stock_id, stock_name, shares, weight}.
    """
    domain = _get_domain(CAPITALFUND_URL)
    session = get_session()
    _wait_for_rate_limit(domain)
    try:
        resp = session.post(
            CAPITALFUND_URL,
            json={"fundId": fund_id},
            timeout=30,
        )
        resp.raise_for_status()
        data = resp.json()
    except Exception as e:
        print(f"  [ERROR] Failed to fetch capitalfund API for fundId={fund_id}: {e}")
        return []

    stocks = data.get("data", data).get("stocks", [])
    holdings = []
    for s in stocks:
        stock_id = (s.get("stocNo") or "").strip()
        stock_name = (s.get("stocName") or "").strip()
        shares = s.get("share")
        weight = s.get("weight")
        if stock_id and shares is not None:
            holdings.append({
                "stock_id": stock_id,
                "stock_name": stock_name,
                "shares": int(shares),
                "weight": float(weight) if weight is not None else None,
            })

    return holdings

# ---------------------------------------------------------------------------
# Fhtrust (復華投信) — Excel download
# ---------------------------------------------------------------------------

FHTRUST_EXCEL_URL = "https://www.fhtrust.com.tw/api/assetsExcel"


def _fetch_fhtrust(etf_code: str, trade_date: date) -> list[dict]:
    """
    Fetch holdings from fhtrust Excel download.
    Returns list of {stock_id, stock_name, shares, weight}.
    """
    from openpyxl import load_workbook
    from io import BytesIO

    date_str = trade_date.strftime("%Y%m%d")
    url = f"{FHTRUST_EXCEL_URL}/{etf_code}/{date_str}"
    resp = fetch(url)
    if resp is None or resp.status_code != 200:
        print(f"  [ERROR] Failed to fetch fhtrust Excel for {etf_code}/{date_str}")
        return []

    wb = load_workbook(BytesIO(resp.content))
    ws = wb.active

    holdings = []
    for row in ws.iter_rows(min_row=1, values_only=True):
        if not row[0]:
            continue
        cell0 = str(row[0]).strip()
        # Stock rows: 4-digit numeric code
        if cell0.isdigit() and len(cell0) == 4:
            stock_id = cell0
            stock_name = str(row[1] or "").strip()
            shares_str = str(row[2] or "0").replace(",", "")
            weight_str = str(row[4] or "0").replace("%", "").replace(",", "")
            try:
                shares = int(shares_str)
                weight = float(weight_str)
            except ValueError:
                continue
            if shares > 0:
                holdings.append({
                    "stock_id": stock_id,
                    "stock_name": stock_name,
                    "shares": shares,
                    "weight": weight,
                })

    return holdings


# ---------------------------------------------------------------------------
# Nomura (野村投信) — JSON API
# ---------------------------------------------------------------------------

NOMURA_API_URL = "https://www.nomurafunds.com.tw/API/ETFAPI/api/Fund/GetFundAssets"


def _fetch_nomura(fund_no: str, trade_date: date) -> list[dict]:
    """
    Fetch holdings from Nomura ETFWEB API.
    Returns list of {stock_id, stock_name, shares, weight}.
    """
    domain = _get_domain(NOMURA_API_URL)
    session = get_session()
    _wait_for_rate_limit(domain)
    try:
        resp = session.post(
            NOMURA_API_URL,
            json={
                "FundID": fund_no,
                "SearchDate": trade_date.strftime("%Y-%m-%dT00:00:00"),
            },
            timeout=30,
        )
        resp.raise_for_status()
        data = resp.json()
    except Exception as e:
        print(f"  [ERROR] Nomura API failed for {fund_no}: {e}")
        return []

    entries = data.get("Entries") or {}
    table_data = (entries.get("Data") or {}).get("Table") or []

    holdings = []
    for table in table_data:
        rows = table.get("Rows") or []
        cols = table.get("Columns") or []
        # Stock table has columns: 證券代號, 證券名稱, 股數, 比重(%)
        if len(cols) < 4:
            continue
        for row in rows:
            if len(row) < 4:
                continue
            stock_id = str(row[0]).strip()
            if not stock_id or not stock_id.isdigit() or len(stock_id) != 4:
                continue
            stock_name = str(row[1]).strip()
            try:
                shares = int(str(row[2]).replace(",", ""))
                weight = float(str(row[3]).replace(",", ""))
            except (ValueError, TypeError):
                continue
            if shares > 0:
                holdings.append({
                    "stock_id": stock_id,
                    "stock_name": stock_name,
                    "shares": shares,
                    "weight": weight,
                })

    return holdings


# ---------------------------------------------------------------------------
# Allianz (安聯投信) — JSON API with XSRF token
# ---------------------------------------------------------------------------

ALLIANZ_API_BASE = "https://etf.allianzgi.com.tw/webapi/api"


def _fetch_allianz(fund_id: str) -> list[dict]:
    """
    Fetch holdings from Allianz ETF API (requires XSRF token).
    Returns list of {stock_id, stock_name, shares, weight}.
    """
    session = get_session()
    headers = {"User-Agent": "Mozilla/5.0"}

    # Step 1: obtain XSRF token
    _wait_for_rate_limit(_get_domain(ALLIANZ_API_BASE))
    try:
        token_resp = session.get(
            f"{ALLIANZ_API_BASE}/AntiForgery/GetAntiForgeryToken",
            headers=headers,
            timeout=15,
        )
        token_resp.raise_for_status()
        xsrf_token = token_resp.json()["token"]
    except Exception as e:
        print(f"  [ERROR] Allianz XSRF token failed: {e}")
        return []

    # Step 2: fetch holdings
    _wait_for_rate_limit(_get_domain(ALLIANZ_API_BASE))
    try:
        resp = session.post(
            f"{ALLIANZ_API_BASE}/Fund/GetFundAssets",
            json={"FundID": fund_id},
            headers={
                **headers,
                "Content-Type": "application/json",
                "X-XSRF-TOKEN": xsrf_token,
            },
            timeout=30,
        )
        resp.raise_for_status()
        data = resp.json()
    except Exception as e:
        print(f"  [ERROR] Allianz GetFundAssets failed for {fund_id}: {e}")
        return []

    entries = data.get("Entries") or {}
    table_data = (entries.get("Data") or {}).get("Table") or []

    holdings = []
    for table in table_data:
        rows = table.get("Rows") or []
        cols = table.get("Columns") or []
        # Stock table has columns: 序號, 證券代號, 證券名稱, 股數, 比重(%)
        if len(cols) < 5:
            continue
        for row in rows:
            if len(row) < 5:
                continue
            stock_id = str(row[1]).strip()
            if not stock_id or not stock_id.isdigit() or len(stock_id) != 4:
                continue
            stock_name = str(row[2]).strip()
            try:
                shares = int(str(row[3]).replace(",", ""))
                weight = float(str(row[4]).replace("%", "").replace(",", ""))
            except (ValueError, TypeError):
                continue
            if shares > 0:
                holdings.append({
                    "stock_id": stock_id,
                    "stock_name": stock_name,
                    "shares": shares,
                    "weight": weight,
                })

    return holdings


# ---------------------------------------------------------------------------
# Fetch dispatcher
# ---------------------------------------------------------------------------


def _fetch_holdings(etf: dict, trade_date: date = None) -> list[dict]:
    source = etf["source"]
    if source == "ezmoney":
        return _fetch_ezmoney(etf["fund_code"])
    elif source == "capitalfund":
        return _fetch_capitalfund(etf["fund_id"])
    elif source == "fhtrust":
        if trade_date is None:
            trade_date = date.today()
        return _fetch_fhtrust(etf["etf_code"], trade_date)
    elif source == "nomura":
        if trade_date is None:
            trade_date = date.today()
        return _fetch_nomura(etf["fund_no"], trade_date)
    elif source == "allianz":
        return _fetch_allianz(etf["fund_id"])
    else:
        print(f"  [ERROR] Unknown source: {source}")
        return []

# ---------------------------------------------------------------------------
# Diff computation
# ---------------------------------------------------------------------------


def _compute_diff(prev: list[dict], curr: list[dict]) -> list[dict]:
    """Compare two holdings snapshots, return list of changes."""
    prev_map = {h["stock_id"]: h for h in prev}
    curr_map = {h["stock_id"]: h for h in curr}
    all_ids = set(prev_map) | set(curr_map)

    diffs = []
    for sid in all_ids:
        p = prev_map.get(sid)
        c = curr_map.get(sid)

        if p and not c:
            diffs.append({
                "stock_id": sid,
                "stock_name": p["stock_name"],
                "change_type": "removed",
                "prev_shares": p["shares"],
                "curr_shares": None,
                "share_diff": None,
                "prev_weight": p["weight"],
                "curr_weight": None,
                "weight_diff": None,
            })
        elif c and not p:
            diffs.append({
                "stock_id": sid,
                "stock_name": c["stock_name"],
                "change_type": "added",
                "prev_shares": None,
                "curr_shares": c["shares"],
                "share_diff": None,
                "prev_weight": None,
                "curr_weight": c["weight"],
                "weight_diff": None,
            })
        elif c["shares"] != p["shares"]:
            share_diff = c["shares"] - p["shares"]
            w_diff = None
            if c["weight"] is not None and p["weight"] is not None:
                w_diff = round(c["weight"] - p["weight"], 4)
            change_type = "increased" if share_diff > 0 else "decreased"
            diffs.append({
                "stock_id": sid,
                "stock_name": c["stock_name"],
                "change_type": change_type,
                "prev_shares": p["shares"],
                "curr_shares": c["shares"],
                "share_diff": share_diff,
                "prev_weight": p["weight"],
                "curr_weight": c["weight"],
                "weight_diff": w_diff,
            })

    return diffs

# ---------------------------------------------------------------------------
# DB operations
# ---------------------------------------------------------------------------


def _get_prev_holdings(cur, etf_id: str, trade_date: date) -> list[dict]:
    """Get the most recent holdings before trade_date."""
    cur.execute("""
        SELECT trade_date, stock_id, stock_name, shares, weight
        FROM tw.etf_holdings
        WHERE etf_id = %s AND trade_date < %s
        ORDER BY trade_date DESC
        LIMIT 200
    """, (etf_id, trade_date))
    rows = cur.fetchall()
    if not rows:
        return []
    # All rows share the same trade_date (the most recent one)
    target_date = rows[0]["trade_date"] if rows else None
    return [
        {
            "stock_id": r["stock_id"],
            "stock_name": r["stock_name"],
            "shares": r["shares"],
            "weight": float(r["weight"]) if r["weight"] is not None else None,
        }
        for r in rows if r["trade_date"] == target_date
    ]


def _save_holdings(cur, etf_id: str, trade_date: date, holdings: list[dict]):
    for h in holdings:
        cur.execute("""
            INSERT INTO tw.etf_holdings (etf_id, trade_date, stock_id, stock_name, shares, weight)
            VALUES (%s, %s, %s, %s, %s, %s)
            ON CONFLICT (etf_id, trade_date, stock_id) DO UPDATE SET
                stock_name = EXCLUDED.stock_name,
                shares = EXCLUDED.shares,
                weight = EXCLUDED.weight
        """, (etf_id, trade_date, h["stock_id"], h["stock_name"], h["shares"], h["weight"]))


def _save_diff(cur, etf_id: str, trade_date: date, diffs: list[dict]):
    for d in diffs:
        cur.execute("""
            INSERT INTO tw.etf_holdings_diff
                (etf_id, trade_date, stock_id, stock_name, change_type,
                 prev_shares, curr_shares, share_diff, prev_weight, curr_weight, weight_diff)
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
            ON CONFLICT (etf_id, trade_date, stock_id) DO UPDATE SET
                stock_name = EXCLUDED.stock_name,
                change_type = EXCLUDED.change_type,
                prev_shares = EXCLUDED.prev_shares,
                curr_shares = EXCLUDED.curr_shares,
                share_diff = EXCLUDED.share_diff,
                prev_weight = EXCLUDED.prev_weight,
                curr_weight = EXCLUDED.curr_weight,
                weight_diff = EXCLUDED.weight_diff
        """, (
            etf_id, trade_date, d["stock_id"], d["stock_name"], d["change_type"],
            d["prev_shares"], d["curr_shares"], d["share_diff"],
            d["prev_weight"], d["curr_weight"], d["weight_diff"],
        ))

# ---------------------------------------------------------------------------
# Main entry point
# ---------------------------------------------------------------------------


_CHANGE_ORDER = {"added": 0, "removed": 1, "increased": 2, "decreased": 3}
_CHANGE_LABEL = {"added": "新增", "removed": "移除", "increased": "加碼", "decreased": "減碼"}

_LARGE_WEIGHT_THRESHOLD = 0.3   # |weight_diff| >= this => 大部位
_CONSECUTIVE_DAYS = 3           # streak length to flag

# Taiwan convention: red = up/buy, green = down/sell
_ROW_FILLS = {
    ("added",     False): "FFC7CE",  # light red (rare — added is always flagged)
    ("added",     True):  "FF6B6B",  # deep red
    ("increased", False): "FFE4E4",  # very light red
    ("increased", True):  "FF9999",  # medium red
    ("removed",   False): "C6EFCE",  # light green (rare)
    ("removed",   True):  "6BCB77",  # deep green
    ("decreased", False): "E4F5E4",  # very light green
    ("decreased", True):  "99DD99",  # medium green
}


def _is_buy(ct):   return ct in ("added", "increased")
def _is_sell(ct):  return ct in ("removed", "decreased")


def export_recent_diffs_excel(trade_date: date, days: int = 5):
    """Export last N trading days of diffs to a single Excel with one sheet per day."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    with get_cursor() as cur:
        cur.execute("""
            SELECT DISTINCT trade_date
            FROM tw.etf_holdings_diff
            WHERE trade_date <= %s
            ORDER BY trade_date DESC
            LIMIT %s
        """, (trade_date, days))
        dates = [r["trade_date"] for r in cur.fetchall()]

        if not dates:
            print("  [WARN] No diff data to export.")
            return

        # Load extended history for consecutive-streak lookup
        earliest = dates[-1]
        cur.execute("""
            SELECT trade_date, etf_id, stock_id, change_type
            FROM tw.etf_holdings_diff
            WHERE trade_date <= %s
            ORDER BY trade_date DESC
        """, (trade_date,))
        history = cur.fetchall()
        # history_by_key: (etf_id, stock_id) -> list of (trade_date, change_type) desc
        history_by_key = {}
        for h in history:
            history_by_key.setdefault((h["etf_id"], h["stock_id"]), []).append(
                (h["trade_date"], h["change_type"])
            )

        def streak_on(d, etf_id, stock_id, direction_fn):
            """Count consecutive trade dates up to and including d with same direction."""
            lst = history_by_key.get((etf_id, stock_id), [])
            # lst is desc; find entries at d or before with same direction, contiguous in lst order
            streak = 0
            started = False
            for td, ct in lst:
                if td > d:
                    continue
                if not started:
                    if td != d:
                        return 0
                    started = True
                if direction_fn(ct):
                    streak += 1
                else:
                    break
            return streak

        wb = Workbook()
        wb.remove(wb.active)

        header = ["ETF代號", "變動類型", "股票代號", "股票名稱",
                  "張數變動", "前日權重", "當日權重", "權重變動", "備註"]
        header_font = Font(bold=True)
        header_fill = PatternFill("solid", fgColor="DDDDDD")

        for d in dates:
            cur.execute("""
                SELECT etf_id, change_type, stock_id, stock_name,
                       prev_shares, curr_shares, share_diff,
                       prev_weight, curr_weight, weight_diff
                FROM tw.etf_holdings_diff
                WHERE trade_date = %s
            """, (d,))
            rows = cur.fetchall()

            ws = wb.create_sheet(title=d.strftime("%m-%d"))
            ws.append(header)
            for cell in ws[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center")

            def effective_share_diff(r):
                if r["change_type"] == "removed":
                    return -(r["prev_shares"] or 0)
                if r["change_type"] == "added":
                    return r["curr_shares"] or 0
                return r["share_diff"] or 0

            rows_sorted = sorted(
                rows,
                key=lambda r: (
                    r["etf_id"],
                    _CHANGE_ORDER.get(r["change_type"], 9),
                    -abs(effective_share_diff(r)),
                ),
            )

            prev_etf = None
            for r in rows_sorted:
                if prev_etf is not None and r["etf_id"] != prev_etf:
                    ws.append([])
                prev_etf = r["etf_id"]

                ct = r["change_type"]
                wdiff = float(r["weight_diff"]) if r["weight_diff"] is not None else None
                is_large = (wdiff is not None and abs(wdiff) >= _LARGE_WEIGHT_THRESHOLD) \
                           or ct in ("added", "removed")

                notes = []
                if is_large and ct in ("increased", "decreased"):
                    notes.append("大部位")
                direction_fn = _is_buy if _is_buy(ct) else _is_sell
                streak = streak_on(d, r["etf_id"], r["stock_id"], direction_fn)
                if streak >= _CONSECUTIVE_DAYS:
                    verb = "加碼" if _is_buy(ct) else "減碼"
                    notes.append(f"連續{verb}{streak}天")

                ws.append([
                    r["etf_id"],
                    _CHANGE_LABEL.get(ct, ct),
                    r["stock_id"], r["stock_name"],
                    effective_share_diff(r),
                    float(r["prev_weight"]) if r["prev_weight"] is not None else None,
                    float(r["curr_weight"]) if r["curr_weight"] is not None else None,
                    wdiff,
                    "｜".join(notes),
                ])

                fill_color = _ROW_FILLS.get((ct, is_large))
                if fill_color:
                    fill = PatternFill("solid", fgColor=fill_color)
                    for cell in ws[ws.max_row]:
                        cell.fill = fill

            widths = [10, 11, 10, 36, 14, 12, 12, 12, 22]
            for i, w in enumerate(widths, 1):
                ws.column_dimensions[chr(64 + i)].width = w
            ws.freeze_panes = "A2"

    os.makedirs(REPORT_DIR, exist_ok=True)
    out_path = os.path.join(REPORT_DIR, f"active_etf_diff_{trade_date.strftime('%Y%m%d')}.xlsx")
    wb.save(out_path)
    print(f"  Excel exported: {out_path} ({len(dates)} sheets)")


def scrape_date(trade_date: date):
    """Scrape ETF holdings for all tracked ETFs and compute diffs."""
    for etf in ETF_REGISTRY:
        etf_id = etf["etf_id"]
        print(f"  Fetching holdings for {etf_id} ...")

        holdings = _fetch_holdings(etf, trade_date)
        if not holdings:
            print(f"  [WARN] No holdings data for {etf_id}, skipping.")
            continue

        print(f"  {etf_id}: {len(holdings)} stocks")

        with get_cursor() as cur:
            prev = _get_prev_holdings(cur, etf_id, trade_date)
            _save_holdings(cur, etf_id, trade_date, holdings)

            if prev:
                diffs = _compute_diff(prev, holdings)
                _save_diff(cur, etf_id, trade_date, diffs)

                added = sum(1 for d in diffs if d["change_type"] == "added")
                removed = sum(1 for d in diffs if d["change_type"] == "removed")
                changed = len(diffs) - added - removed
                print(f"  {etf_id} diff: +{added} added, -{removed} removed, ~{changed} changed")
            else:
                print(f"  {etf_id}: first snapshot, no diff computed.")

    export_recent_diffs_excel(trade_date)
