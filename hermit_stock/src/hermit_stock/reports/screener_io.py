"""Output writers for screener results: Excel (openpyxl) + markdown table."""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from .screener import ScreenRow

RULE_CODES = ("F1", "F2", "F3", "F4", "F5", "F6", "F7", "F8")


def _rule_mark(passed: bool | None) -> str:
    if passed is True:
        return "✓"
    if passed is False:
        return "✗"
    return "?"


def _row_dict(r: ScreenRow) -> dict[str, object]:
    by_code = {rr.code: rr for rr in r.rule_results}
    out: dict[str, object] = {
        "ticker": r.ticker,
        "name": r.name,
        "industry": r.industry or "",
    }
    for code in RULE_CODES:
        out[code] = _rule_mark(by_code[code].passed) if code in by_code else "—"
    out["total"] = r.scoreboard.score
    out["grade"] = r.scoreboard.grade
    if r.valuation is None:
        out["method"] = "—"
        out["multiple"] = ""
        out["band"] = "—"
        out["upside"] = ""
    else:
        v = r.valuation
        out["method"] = v.method
        out["multiple"] = round(v.current_multiple, 2) if v.current_multiple is not None else ""
        out["band"] = v.band_position or "—"
        out["upside"] = round(v.upside_mean * 100, 2) if v.upside_mean is not None else ""
    return out


COLUMNS = [
    "ticker",
    "name",
    "industry",
    *RULE_CODES,
    "total",
    "grade",
    "method",
    "multiple",
    "band",
    "upside",
]


def to_excel(rows: list[ScreenRow], path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "screener"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="305496")
    grade_fills = {
        "A": PatternFill("solid", fgColor="C6EFCE"),
        "B": PatternFill("solid", fgColor="FFEB9C"),
        "C": PatternFill("solid", fgColor="FFD7B5"),
        "D": PatternFill("solid", fgColor="FFC7CE"),
    }

    ws.append(COLUMNS)
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill

    for r in rows:
        d = _row_dict(r)
        ws.append([d[c] for c in COLUMNS])

    grade_col = COLUMNS.index("grade") + 1
    for row in ws.iter_rows(min_row=2, min_col=grade_col, max_col=grade_col):
        for cell in row:
            fill = grade_fills.get(str(cell.value))
            if fill is not None:
                cell.fill = fill

    widths = {
        "ticker": 8,
        "name": 14,
        "industry": 12,
        "total": 6,
        "grade": 6,
        "method": 8,
        "multiple": 9,
        "band": 14,
        "upside": 9,
    }
    for col_idx, key in enumerate(COLUMNS, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = widths.get(key, 6)

    ws.freeze_panes = "D2"  # freeze ticker/name/industry + header
    wb.save(path)


def to_markdown(rows: list[ScreenRow], top_n: int | None = None) -> str:
    lines: list[str] = []
    head = "| " + " | ".join(COLUMNS) + " |"
    sep = "|" + "|".join(["---"] * len(COLUMNS)) + "|"
    lines.append(head)
    lines.append(sep)
    cut = rows if top_n is None else rows[:top_n]
    for r in cut:
        d = _row_dict(r)
        lines.append("| " + " | ".join(str(d[c]) for c in COLUMNS) + " |")
    return "\n".join(lines)
