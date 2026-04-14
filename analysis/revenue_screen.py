"""
Unified revenue screener — runs all strategies and exports to a single Excel
with one sheet per strategy.

Usage:
  python -m analysis.revenue_screen                # latest month
  python -m analysis.revenue_screen 2026-03        # specific month
"""

import sys
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill

from db.connection import get_cursor
from analysis.revenue_three_arrows import screen as three_arrows_screen
from analysis.revenue_turnaround import screen as turnaround_screen
from analysis.revenue_streak import screen as streak_screen
from analysis.revenue_off_season import screen as off_season_screen
from analysis.revenue_deceleration import screen as decel_screen
from analysis.revenue_decline import screen as decline_screen
from analysis.revenue_historic_low import screen as historic_low_screen
from analysis.revenue_peak_miss import screen as peak_miss_screen


def _write_sheet(wb, title: str, color: str, headers: list[tuple],
                 rows: list[dict], columns: list[tuple],
                 tab_color: str = None):
    """
    Generic helper to populate one worksheet.
    columns: list of (dict_key, number_format or None)
    """
    ws = wb.create_sheet(title=title)
    if tab_color:
        ws.sheet_properties.tabColor = tab_color

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for col, (h_title, width) in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h_title)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        ws.column_dimensions[cell.column_letter].width = width

    for row_idx, r in enumerate(rows, 2):
        for col, (key, fmt) in enumerate(columns, 1):
            val = r.get(key)
            if val is None:
                val = ""
            cell = ws.cell(row=row_idx, column=col, value=val)
            if fmt:
                cell.number_format = fmt

    ws.freeze_panes = "A2"
    if rows:
        ws.auto_filter.ref = ws.dimensions


def _build_elite(strategy_map: dict[str, list[dict]],
                  sort_ascending: bool = False) -> list[dict]:
    """
    Build elite picks: stocks appearing in 2+ strategies.
    strategy_map: {"策略名": [result_dicts, ...], ...}
    sort_ascending: True for bearish (worst YoY first).
    """
    indexed = {name: {r["stock_id"]: r for r in rows}
               for name, rows in strategy_map.items()}

    all_ids = set()
    for stocks in indexed.values():
        all_ids.update(stocks.keys())

    results = []
    for sid in all_ids:
        matched = [name for name, stocks in indexed.items() if sid in stocks]
        if len(matched) < 2:
            continue

        base = None
        for stocks in indexed.values():
            if sid in stocks:
                base = stocks[sid]
                break

        results.append({
            "stock_id": sid,
            "name": base["name"],
            "market": base["market"],
            "industry": base.get("industry"),
            "revenue": base["revenue"],
            "yoy_pct": base.get("yoy_pct"),
            "mom_pct": base.get("mom_pct"),
            "hit_count": len(matched),
            "strategies": "、".join(matched),
        })

    if sort_ascending:
        results.sort(key=lambda x: (-x["hit_count"], (x["yoy_pct"] or 0)))
    else:
        results.sort(key=lambda x: (-x["hit_count"], -(x["yoy_pct"] or 0)))
    return results


def export_all(year_month: str):
    print(f"\n=== Revenue Screener: {year_month} ===\n")

    arrows = three_arrows_screen(year_month)
    turnaround = turnaround_screen(year_month)
    streak = streak_screen(year_month)
    decel = decel_screen(year_month)
    decline = decline_screen(year_month)
    historic_low = historic_low_screen(year_month)
    peak_miss = peak_miss_screen(year_month)

    off_season = off_season_screen(year_month)

    # Elite picks
    elite_long = _build_elite({
        "三支箭": arrows,
        "連續成長": streak,
        "轉機": turnaround,
        "淡季不淡": off_season,
    })
    elite_short = _build_elite({
        "成長減速": decel,
        "連續衰退": decline,
        "歷史新低": historic_low,
        "旺季不旺": peak_miss,
    }, sort_ascending=True)

    path = f"revenue_screen_{year_month}.xlsx"
    wb = Workbook()
    wb.remove(wb.active)

    TAB_RED = "FF0000"    # bullish
    TAB_GREEN = "00B050"  # bearish

    elite_headers = [
        ("代號", 10), ("名稱", 14), ("市場", 8), ("產業", 14),
        ("當月營收(千)", 16), ("YoY%", 12), ("MoM%", 10),
        ("命中數", 8), ("命中策略", 30)]
    elite_columns = [
        ("stock_id", None), ("name", None), ("market", None), ("industry", None),
        ("revenue", "#,##0"), ("yoy_pct", "#,##0.00"), ("mom_pct", "#,##0.00"),
        ("hit_count", None), ("strategies", None)]

    # --- Sheet 1: Long Elite ---
    _write_sheet(wb, "做多精選", "002060",
        elite_headers, elite_long, elite_columns,
        tab_color=TAB_RED)

    # --- Sheet 2: Short Elite ---
    _write_sheet(wb, "做空精選", "002060",
        elite_headers, elite_short, elite_columns,
        tab_color=TAB_GREEN)

    # --- Sheet 2: Three Arrows ---
    _write_sheet(wb, "營收三支箭", "4472C4",
        [("代號", 10), ("名稱", 14), ("市場", 8), ("產業", 14),
         ("當月營收(千)", 16), ("歷史平均(千)", 16), ("超越均值%", 12),
         ("單月YoY%", 12), ("累計YoY%", 12), ("前歷史高(千)", 16),
         ("MoM%", 10), ("備註", 30)],
        arrows,
        [("stock_id", None), ("name", None), ("market", None), ("industry", None),
         ("revenue", "#,##0"), ("avg_revenue", "#,##0"), ("rev_vs_avg_pct", "#,##0.00"),
         ("yoy_pct", "#,##0.00"), ("cum_yoy_pct", "#,##0.00"),
         ("prev_max_revenue", "#,##0"), ("mom_pct", "#,##0.00"), ("note", None)],
        tab_color=TAB_RED)

    # --- Sheet 3: Turnaround ---
    _write_sheet(wb, "營收轉機", "ED7D31",
        [("代號", 10), ("名稱", 14), ("市場", 8), ("產業", 14),
         ("當月營收(千)", 16), ("當月YoY%", 12), ("前月YoY%", 12),
         ("連續衰退月數", 14), ("MoM%", 10), ("備註", 30)],
        turnaround,
        [("stock_id", None), ("name", None), ("market", None), ("industry", None),
         ("revenue", "#,##0"), ("yoy_pct", "#,##0.00"), ("prev_yoy_pct", "#,##0.00"),
         ("decline_months", None), ("mom_pct", "#,##0.00"), ("note", None)],
        tab_color=TAB_RED)

    # --- Sheet 4: Consecutive Growth ---
    _write_sheet(wb, "營收連續成長", "70AD47",
        [("代號", 10), ("名稱", 14), ("市場", 8), ("產業", 14),
         ("當月營收(千)", 16), ("當月YoY%", 12), ("連續成長月數", 14),
         ("期間平均YoY%", 14), ("期間最低YoY%", 14), ("期間最高YoY%", 14),
         ("MoM%", 10), ("備註", 30)],
        streak,
        [("stock_id", None), ("name", None), ("market", None), ("industry", None),
         ("revenue", "#,##0"), ("yoy_pct", "#,##0.00"), ("streak", None),
         ("avg_yoy_pct", "#,##0.00"), ("min_yoy_pct", "#,##0.00"),
         ("max_yoy_pct", "#,##0.00"), ("mom_pct", "#,##0.00"), ("note", None)],
        tab_color=TAB_RED)

    # --- Sheet 5: Deceleration Warning ---
    _write_sheet(wb, "成長減速預警", "C00000",
        [("代號", 10), ("名稱", 14), ("市場", 8), ("產業", 14),
         ("當月營收(千)", 16), ("當月YoY%", 12), ("波段最高YoY%", 14),
         ("YoY下降幅度", 14), ("連續成長月數", 14), ("連續減速月數", 14),
         ("MoM%", 10), ("備註", 30)],
        decel,
        [("stock_id", None), ("name", None), ("market", None), ("industry", None),
         ("revenue", "#,##0"), ("yoy_pct", "#,##0.00"), ("peak_yoy_pct", "#,##0.00"),
         ("yoy_drop_pct", "#,##0.00"), ("growth_streak", None),
         ("decel_months", None), ("mom_pct", "#,##0.00"), ("note", None)],
        tab_color=TAB_GREEN)

    # --- Sheet 6: Consecutive Decline ---
    _write_sheet(wb, "營收連續衰退", "C00000",
        [("代號", 10), ("名稱", 14), ("市場", 8), ("產業", 14),
         ("當月營收(千)", 16), ("當月YoY%", 12), ("連續衰退月數", 14),
         ("期間平均YoY%", 14), ("期間最差YoY%", 14), ("衰退加深", 10),
         ("MoM%", 10), ("備註", 30)],
        decline,
        [("stock_id", None), ("name", None), ("market", None), ("industry", None),
         ("revenue", "#,##0"), ("yoy_pct", "#,##0.00"), ("streak", None),
         ("avg_yoy_pct", "#,##0.00"), ("worst_yoy_pct", "#,##0.00"),
         ("deepening", None), ("mom_pct", "#,##0.00"), ("note", None)],
        tab_color=TAB_GREEN)

    # --- Sheet 7: Historic Low ---
    _write_sheet(wb, "營收歷史新低", "C00000",
        [("代號", 10), ("名稱", 14), ("市場", 8), ("產業", 14),
         ("當月營收(千)", 16), ("前歷史低(千)", 16), ("歷史平均(千)", 16),
         ("低於均值%", 12), ("YoY%", 10), ("MoM%", 10), ("備註", 30)],
        historic_low,
        [("stock_id", None), ("name", None), ("market", None), ("industry", None),
         ("revenue", "#,##0"), ("prev_min_revenue", "#,##0"), ("avg_revenue", "#,##0"),
         ("below_avg_pct", "#,##0.00"), ("yoy_pct", "#,##0.00"),
         ("mom_pct", "#,##0.00"), ("note", None)],
        tab_color=TAB_GREEN)

    # --- Sheet 8: Peak Miss ---
    _write_sheet(wb, "旺季不旺", "C00000",
        [("代號", 10), ("名稱", 14), ("市場", 8), ("產業", 14),
         ("當月營收(千)", 16), ("月份", 8), ("旺季係數", 10),
         ("季節CV", 10), ("歷年同月均值(千)", 18), ("落後均值%", 12),
         ("YoY%", 10), ("MoM%", 10), ("備註", 30)],
        peak_miss,
        [("stock_id", None), ("name", None), ("market", None), ("industry", None),
         ("revenue", "#,##0"), ("period", None), ("seasonal_coeff", "0.000"),
         ("cv", "0.0000"), ("hist_avg", "#,##0"), ("miss_pct", "#,##0.00"),
         ("yoy_pct", "#,##0.00"), ("mom_pct", "#,##0.00"), ("note", None)],
        tab_color=TAB_GREEN)

    # --- Sheet: Off-Season ---
    _write_sheet(wb, "淡季不淡", "7030A0",
        [("代號", 10), ("名稱", 14), ("市場", 8), ("產業", 14),
         ("當月營收(千)", 16), ("月份", 8), ("淡季係數", 10),
         ("季節CV", 10), ("歷年同月均值(千)", 18), ("超越均值%", 12),
         ("YoY%", 10), ("MoM%", 10), ("備註", 30)],
        off_season,
        [("stock_id", None), ("name", None), ("market", None), ("industry", None),
         ("revenue", "#,##0"), ("period", None), ("seasonal_coeff", "0.000"),
         ("cv", "0.0000"), ("hist_avg", "#,##0"), ("beat_pct", "#,##0.00"),
         ("yoy_pct", "#,##0.00"), ("mom_pct", "#,##0.00"), ("note", None)],
        tab_color=TAB_RED)

    wb.save(path)
    print(f"\nExported to {path}")
    print(f"  做多精選      : {len(elite_long)} stocks")
    print(f"  做空精選      : {len(elite_short)} stocks")
    print(f"  營收三支箭    : {len(arrows)} stocks")
    print(f"  營收轉機      : {len(turnaround)} stocks")
    print(f"  營收連續成長  : {len(streak)} stocks")
    print(f"  成長減速預警  : {len(decel)} stocks")
    print(f"  營收連續衰退  : {len(decline)} stocks")
    print(f"  營收歷史新低  : {len(historic_low)} stocks")
    print(f"  旺季不旺      : {len(peak_miss)} stocks")
    print(f"  淡季不淡      : {len(off_season)} stocks")
    return path


if __name__ == "__main__":
    ym = sys.argv[1] if len(sys.argv) >= 2 else None
    if not ym:
        with get_cursor(commit=False) as cur:
            cur.execute("SELECT MAX(year_month) AS ym FROM tw.monthly_revenue")
            ym = cur.fetchone()["ym"]
    export_all(ym)
