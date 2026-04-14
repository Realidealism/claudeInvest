"""
Revenue growth deceleration warning (營收成長減速預警).

Target: stocks with consecutive YoY growth, but YoY% has been
declining for N or more recent months. Growth is still positive
but momentum is fading — potential peak signal.

Usage:
  python -m analysis.revenue_deceleration                # latest month
  python -m analysis.revenue_deceleration 2026-03        # specific month
"""

import sys
from db.connection import get_cursor

# Minimum consecutive months of positive YoY to qualify
MIN_GROWTH_STREAK = 6

# Minimum consecutive months of declining YoY% (while still positive)
MIN_DECEL_MONTHS = 3


def screen(year_month: str = None) -> list[dict]:
    with get_cursor(commit=False) as cur:
        if not year_month:
            cur.execute("SELECT MAX(year_month) AS ym FROM tw.monthly_revenue")
            year_month = cur.fetchone()["ym"]

        print(f"Deceleration screening for {year_month} ...")

        # Get stocks with positive YoY in target month
        cur.execute("""
            SELECT r.stock_id, r.revenue, r.yoy_pct, r.mom_pct, r.note,
                   s.name, s.market, s.industry
            FROM tw.monthly_revenue r
            JOIN tw.stocks s ON r.stock_id = s.stock_id
            WHERE r.year_month = %s
              AND r.yoy_pct > 0
              AND s.is_active = TRUE
        """, (year_month,))
        candidates = cur.fetchall()

        results = []
        for c in candidates:
            stock_id = c["stock_id"]

            # Get recent YoY history descending
            cur.execute("""
                SELECT year_month, yoy_pct
                FROM tw.monthly_revenue
                WHERE stock_id = %s AND year_month <= %s
                ORDER BY year_month DESC
                LIMIT 60
            """, (stock_id, year_month))
            history = cur.fetchall()

            # Count consecutive positive YoY streak
            streak = 0
            for r in history:
                if r["yoy_pct"] is not None and r["yoy_pct"] > 0:
                    streak += 1
                else:
                    break

            if streak < MIN_GROWTH_STREAK:
                continue

            # Count consecutive months where YoY% is declining
            # (each month's YoY < previous month's YoY, from newest backwards)
            decel = 0
            for i in range(len(history) - 1):
                curr = history[i]["yoy_pct"]
                prev = history[i + 1]["yoy_pct"]
                if curr is None or prev is None:
                    break
                if curr < prev and curr > 0:
                    decel += 1
                else:
                    break

            if decel < MIN_DECEL_MONTHS:
                continue

            # Peak YoY in this growth streak
            peak_yoy = max(
                float(history[i]["yoy_pct"])
                for i in range(streak)
                if history[i]["yoy_pct"] is not None
            )
            current_yoy = float(c["yoy_pct"])

            results.append({
                "stock_id": stock_id,
                "name": c["name"],
                "market": c["market"],
                "industry": c["industry"],
                "revenue": c["revenue"],
                "yoy_pct": current_yoy,
                "peak_yoy_pct": peak_yoy,
                "yoy_drop_pct": round(peak_yoy - current_yoy, 2),
                "growth_streak": streak,
                "decel_months": decel,
                "mom_pct": float(c["mom_pct"]) if c["mom_pct"] is not None else None,
                "note": c["note"],
            })

        results.sort(key=lambda x: x["decel_months"], reverse=True)
        print(f"Found {len(results)} decelerating stocks.")
        return results


def export_excel(results: list[dict], year_month: str, path: str = None):
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill

    if not path:
        path = f"revenue_deceleration_{year_month}.xlsx"

    wb = Workbook()
    ws = wb.active
    ws.title = "Deceleration"

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    headers = [
        ("代號", 10), ("名稱", 14), ("市場", 8), ("產業", 14),
        ("當月營收(千)", 16), ("當月YoY%", 12), ("波段最高YoY%", 14),
        ("YoY下降幅度", 14), ("連續成長月數", 14), ("連續減速月數", 14),
        ("MoM%", 10), ("備註", 30),
    ]

    for col, (title, width) in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=title)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        ws.column_dimensions[cell.column_letter].width = width

    for row_idx, r in enumerate(results, 2):
        ws.cell(row=row_idx, column=1, value=r["stock_id"])
        ws.cell(row=row_idx, column=2, value=r["name"])
        ws.cell(row=row_idx, column=3, value=r["market"])
        ws.cell(row=row_idx, column=4, value=r["industry"])
        ws.cell(row=row_idx, column=5, value=r["revenue"]).number_format = "#,##0"
        ws.cell(row=row_idx, column=6, value=r["yoy_pct"]).number_format = "#,##0.00"
        ws.cell(row=row_idx, column=7, value=r["peak_yoy_pct"]).number_format = "#,##0.00"
        ws.cell(row=row_idx, column=8, value=r["yoy_drop_pct"]).number_format = "#,##0.00"
        ws.cell(row=row_idx, column=9, value=r["growth_streak"])
        ws.cell(row=row_idx, column=10, value=r["decel_months"])
        ws.cell(row=row_idx, column=11, value=r["mom_pct"]).number_format = "#,##0.00"
        ws.cell(row=row_idx, column=12, value=r["note"] or "")

    ws.freeze_panes = "A2"
    if results:
        ws.auto_filter.ref = ws.dimensions
    wb.save(path)
    print(f"Exported to {path}")
    return path


if __name__ == "__main__":
    ym = sys.argv[1] if len(sys.argv) >= 2 else None
    results = screen(ym)
    if results:
        if not ym:
            with get_cursor(commit=False) as cur:
                cur.execute("SELECT MAX(year_month) AS ym FROM tw.monthly_revenue")
                ym = cur.fetchone()["ym"]
        export_excel(results, ym)
    else:
        print("No decelerating stocks found.")
