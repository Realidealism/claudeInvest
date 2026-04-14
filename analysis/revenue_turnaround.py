"""
Revenue turnaround stock screener (營收轉機股).

Condition: YoY was negative for N consecutive months, then turns positive
in the target month. Catches inflection points.

Usage:
  python -m analysis.revenue_turnaround                # latest month
  python -m analysis.revenue_turnaround 2026-03        # specific month
"""

import sys
from db.connection import get_cursor

# Minimum consecutive months of negative YoY before the turnaround
MIN_DECLINE_MONTHS = 3


def _prev_months(year_month: str, n: int) -> list[str]:
    """Return the previous n year_month strings in descending order."""
    y = int(year_month[:4])
    m = int(year_month[5:7])
    result = []
    for _ in range(n):
        m -= 1
        if m == 0:
            m = 12
            y -= 1
        result.append(f"{y}-{m:02d}")
    return result


def screen(year_month: str = None) -> list[dict]:
    with get_cursor(commit=False) as cur:
        if not year_month:
            cur.execute("SELECT MAX(year_month) AS ym FROM tw.monthly_revenue")
            year_month = cur.fetchone()["ym"]

        print(f"Turnaround screening for {year_month} ...")

        prev_months = _prev_months(year_month, MIN_DECLINE_MONTHS)

        # Get stocks with positive YoY in target month
        cur.execute("""
            SELECT r.stock_id, r.revenue, r.yoy_pct, r.mom_pct, r.note,
                   s.name, s.market, s.industry
            FROM tw.monthly_revenue r
            JOIN tw.stocks s ON r.stock_id = s.stock_id
            WHERE r.year_month = %s
              AND r.yoy_pct > 0
              AND s.is_active = TRUE
        """, (year_month,))
        candidates = cur.fetchall()

        results = []
        for c in candidates:
            stock_id = c["stock_id"]

            # Check previous N months all had negative YoY
            cur.execute("""
                SELECT year_month, yoy_pct
                FROM tw.monthly_revenue
                WHERE stock_id = %s AND year_month = ANY(%s)
                ORDER BY year_month DESC
            """, (stock_id, prev_months))
            prev_data = cur.fetchall()

            if len(prev_data) < MIN_DECLINE_MONTHS:
                continue

            all_negative = all(
                r["yoy_pct"] is not None and r["yoy_pct"] < 0
                for r in prev_data
            )
            if not all_negative:
                continue

            # Count total consecutive decline months (beyond minimum)
            decline_count = MIN_DECLINE_MONTHS
            extra_months = _prev_months(prev_months[-1], 24)
            cur.execute("""
                SELECT year_month, yoy_pct
                FROM tw.monthly_revenue
                WHERE stock_id = %s AND year_month = ANY(%s)
                ORDER BY year_month DESC
            """, (stock_id, extra_months))
            for r in cur.fetchall():
                if r["yoy_pct"] is not None and r["yoy_pct"] < 0:
                    decline_count += 1
                else:
                    break

            last_neg_yoy = float(prev_data[0]["yoy_pct"])

            results.append({
                "stock_id": stock_id,
                "name": c["name"],
                "market": c["market"],
                "industry": c["industry"],
                "revenue": c["revenue"],
                "yoy_pct": float(c["yoy_pct"]),
                "prev_yoy_pct": last_neg_yoy,
                "decline_months": decline_count,
                "mom_pct": float(c["mom_pct"]) if c["mom_pct"] is not None else None,
                "note": c["note"],
            })

        results.sort(key=lambda x: x["decline_months"], reverse=True)
        print(f"Found {len(results)} turnaround stocks.")
        return results


def export_excel(results: list[dict], year_month: str, path: str = None):
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill

    if not path:
        path = f"revenue_turnaround_{year_month}.xlsx"

    wb = Workbook()
    ws = wb.active
    ws.title = "Turnaround"

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="ED7D31", end_color="ED7D31", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    headers = [
        ("代號", 10),
        ("名稱", 14),
        ("市場", 8),
        ("產業", 14),
        ("當月營收(千)", 16),
        ("當月YoY%", 12),
        ("前月YoY%", 12),
        ("連續衰退月數", 14),
        ("MoM%", 10),
        ("備註", 30),
    ]

    for col, (title, width) in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=title)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        ws.column_dimensions[cell.column_letter].width = width

    for row_idx, r in enumerate(results, 2):
        ws.cell(row=row_idx, column=1, value=r["stock_id"])
        ws.cell(row=row_idx, column=2, value=r["name"])
        ws.cell(row=row_idx, column=3, value=r["market"])
        ws.cell(row=row_idx, column=4, value=r["industry"])
        ws.cell(row=row_idx, column=5, value=r["revenue"]).number_format = '#,##0'
        ws.cell(row=row_idx, column=6, value=r["yoy_pct"]).number_format = '#,##0.00'
        ws.cell(row=row_idx, column=7, value=r["prev_yoy_pct"]).number_format = '#,##0.00'
        ws.cell(row=row_idx, column=8, value=r["decline_months"])
        ws.cell(row=row_idx, column=9, value=r["mom_pct"]).number_format = '#,##0.00'
        ws.cell(row=row_idx, column=10, value=r["note"] or "")

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    wb.save(path)
    print(f"Exported to {path}")
    return path


if __name__ == "__main__":
    ym = sys.argv[1] if len(sys.argv) >= 2 else None
    results = screen(ym)
    if results:
        if not ym:
            with get_cursor(commit=False) as cur:
                cur.execute("SELECT MAX(year_month) AS ym FROM tw.monthly_revenue")
                ym = cur.fetchone()["ym"]
        export_excel(results, ym)
    else:
        print("No turnaround stocks found.")
