"""
Off-season outperformance stock screener (淡季不淡選股法).

Conditions:
  1. Stock has clear seasonality (CV >= 0.10 after trend removal)
  2. Target month is a "low season" month (coefficient < 1.0)
  3. Revenue beats the historical same-month average

Jan+Feb are merged to neutralize Lunar New Year distortion.

Usage:
  python -m analysis.revenue_off_season                # latest month
  python -m analysis.revenue_off_season 2026-03        # specific month
"""

import sys
from collections import defaultdict
from statistics import mean, stdev

from db.connection import get_cursor

# Minimum CV to qualify as "having clear seasonality"
MIN_CV = 0.15

# Minimum years of history for seasonality calculation
MIN_YEARS = 3


def _compute_seasonality(data: list[tuple]) -> dict | None:
    """
    Compute per-stock seasonality from (year_month, revenue) pairs.
    Returns {period: avg_coefficient, ..., '__cv': cv} or None if insufficient data.
    Periods: '1+2', '3', '4', ..., '12'.
    Jan+Feb revenue is averaged (sum/2) before computing coefficients.
    """
    yearly = defaultdict(lambda: defaultdict(int))
    for ym, rev in data:
        y, m = int(ym[:4]), int(ym[5:7])
        if m in (1, 2):
            yearly[y]["1+2"] += rev
        else:
            yearly[y][str(m)] = rev

    complete_years = [y for y, p in yearly.items() if len(p) >= 11]
    if len(complete_years) < MIN_YEARS:
        return None

    period_coeffs = defaultdict(list)
    for y in complete_years:
        normalized = dict(yearly[y])
        if "1+2" in normalized:
            normalized["1+2"] = normalized["1+2"] / 2
        year_avg = mean(normalized.values())
        if year_avg == 0:
            continue
        for period, rev in normalized.items():
            period_coeffs[period].append(rev / year_avg)

    avg_coeffs = {p: mean(vals) for p, vals in period_coeffs.items()}
    if len(avg_coeffs) < 11:
        return None

    overall_mean = mean(avg_coeffs.values())
    if overall_mean == 0:
        return None

    cv = stdev(avg_coeffs.values()) / overall_mean
    avg_coeffs["__cv"] = cv
    return avg_coeffs


def _same_month_history(cur, stock_id: str, period: str,
                        year_month: str) -> tuple[float | None, int]:
    """
    Get average revenue for the same period across previous years.
    Returns (avg_revenue, year_count).
    """
    target_year = int(year_month[:4])

    if period == "1+2":
        # Sum Jan+Feb per year, then average across years
        cur.execute("""
            SELECT CAST(LEFT(year_month, 4) AS INT) AS y,
                   SUM(revenue) AS total
            FROM tw.monthly_revenue
            WHERE stock_id = %s
              AND RIGHT(year_month, 2) IN ('01', '02')
              AND CAST(LEFT(year_month, 4) AS INT) < %s
            GROUP BY y
            HAVING COUNT(*) = 2
        """, (stock_id, target_year))
        rows = cur.fetchall()
        if not rows:
            return None, 0
        # Divide by 2 to get monthly average for comparison
        avg = mean(r["total"] / 2 for r in rows)
        return avg, len(rows)
    else:
        month_str = period.zfill(2)
        cur.execute("""
            SELECT revenue FROM tw.monthly_revenue
            WHERE stock_id = %s
              AND RIGHT(year_month, 2) = %s
              AND year_month < %s
        """, (stock_id, month_str, year_month))
        rows = cur.fetchall()
        if not rows:
            return None, 0
        avg = mean(r["revenue"] for r in rows)
        return avg, len(rows)


def screen(year_month: str = None, min_cv: float = MIN_CV) -> list[dict]:
    with get_cursor(commit=False) as cur:
        if not year_month:
            cur.execute("SELECT MAX(year_month) AS ym FROM tw.monthly_revenue")
            year_month = cur.fetchone()["ym"]

        print(f"Off-season screening for {year_month} ...")

        target_month = int(year_month[5:7])
        # Determine which period the target month maps to
        if target_month in (1, 2):
            target_period = "1+2"
        else:
            target_period = str(target_month)

        # Load all revenue data for seasonality computation
        cur.execute("""
            SELECT r.stock_id, r.year_month, r.revenue
            FROM tw.monthly_revenue r
            JOIN tw.stocks s ON r.stock_id = s.stock_id
            WHERE s.is_active = TRUE
            ORDER BY r.stock_id, r.year_month
        """)
        all_rows = cur.fetchall()

        stock_data = defaultdict(list)
        for r in all_rows:
            stock_data[r["stock_id"]].append((r["year_month"], r["revenue"]))

        # Get target month revenue and stock info
        cur.execute("""
            SELECT r.stock_id, r.revenue, r.yoy_pct, r.mom_pct, r.note,
                   s.name, s.market, s.industry
            FROM tw.monthly_revenue r
            JOIN tw.stocks s ON r.stock_id = s.stock_id
            WHERE r.year_month = %s AND s.is_active = TRUE
        """, (year_month,))
        target_rows = {r["stock_id"]: r for r in cur.fetchall()}

        # For 1+2 period, also need the other month's revenue
        partner_revenue = {}
        if target_period == "1+2":
            partner_month = 1 if target_month == 2 else 2
            partner_ym = f"{year_month[:4]}-{partner_month:02d}"
            cur.execute("""
                SELECT stock_id, revenue FROM tw.monthly_revenue
                WHERE year_month = %s
            """, (partner_ym,))
            for r in cur.fetchall():
                partner_revenue[r["stock_id"]] = r["revenue"]

        results = []
        seasonal_count = 0

        for stock_id, data in stock_data.items():
            if stock_id not in target_rows:
                continue

            seasonality = _compute_seasonality(data)
            if seasonality is None:
                continue

            cv = seasonality["__cv"]
            if cv < min_cv:
                continue

            seasonal_count += 1
            coeff = seasonality.get(target_period)
            if coeff is None:
                continue

            # Must be a low-season month (coefficient < 1.0)
            if coeff >= 1.0:
                continue

            # Get current period revenue
            c = target_rows[stock_id]
            if target_period == "1+2":
                partner_rev = partner_revenue.get(stock_id)
                if partner_rev is None:
                    continue
                # Monthly average for comparison
                current_rev = float(c["revenue"] + partner_rev) / 2
            else:
                current_rev = float(c["revenue"])

            # Compare against historical same-month average
            hist_avg, year_count = _same_month_history(
                cur, stock_id, target_period, year_month)
            if not hist_avg or year_count < MIN_YEARS:
                continue

            # Must beat historical average
            if current_rev <= hist_avg:
                continue

            beat_pct = round((current_rev - hist_avg) / hist_avg * 100, 2)

            results.append({
                "stock_id": stock_id,
                "name": c["name"],
                "market": c["market"],
                "industry": c["industry"],
                "revenue": c["revenue"],
                "period": target_period,
                "seasonal_coeff": round(coeff, 3),
                "cv": round(cv, 4),
                "hist_avg": round(hist_avg),
                "beat_pct": beat_pct,
                "yoy_pct": float(c["yoy_pct"]) if c["yoy_pct"] is not None else None,
                "mom_pct": float(c["mom_pct"]) if c["mom_pct"] is not None else None,
                "note": c["note"],
            })

        results.sort(key=lambda x: x["beat_pct"], reverse=True)
        print(f"Seasonal stocks (CV >= {min_cv}): {seasonal_count}")
        print(f"Found {len(results)} off-season outperformers.")
        return results


def export_excel(results: list[dict], year_month: str, path: str = None):
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill

    if not path:
        path = f"revenue_off_season_{year_month}.xlsx"

    wb = Workbook()
    ws = wb.active
    ws.title = "Off-Season"

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="7030A0", end_color="7030A0", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    headers = [
        ("代號", 10),
        ("名稱", 14),
        ("市場", 8),
        ("產業", 14),
        ("當月營收(千)", 16),
        ("月份", 8),
        ("淡季係數", 10),
        ("季節CV", 10),
        ("歷年同月均值(千)", 18),
        ("超越均值%", 12),
        ("YoY%", 10),
        ("MoM%", 10),
        ("備註", 30),
    ]

    for col, (title, width) in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=title)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        ws.column_dimensions[cell.column_letter].width = width

    for row_idx, r in enumerate(results, 2):
        ws.cell(row=row_idx, column=1, value=r["stock_id"])
        ws.cell(row=row_idx, column=2, value=r["name"])
        ws.cell(row=row_idx, column=3, value=r["market"])
        ws.cell(row=row_idx, column=4, value=r["industry"])
        ws.cell(row=row_idx, column=5, value=r["revenue"]).number_format = "#,##0"
        ws.cell(row=row_idx, column=6, value=r["period"])
        ws.cell(row=row_idx, column=7, value=r["seasonal_coeff"]).number_format = "0.000"
        ws.cell(row=row_idx, column=8, value=r["cv"]).number_format = "0.0000"
        ws.cell(row=row_idx, column=9, value=r["hist_avg"]).number_format = "#,##0"
        ws.cell(row=row_idx, column=10, value=r["beat_pct"]).number_format = "#,##0.00"
        ws.cell(row=row_idx, column=11, value=r["yoy_pct"]).number_format = "#,##0.00"
        ws.cell(row=row_idx, column=12, value=r["mom_pct"]).number_format = "#,##0.00"
        ws.cell(row=row_idx, column=13, value=r["note"] or "")

    ws.freeze_panes = "A2"
    if results:
        ws.auto_filter.ref = ws.dimensions
    wb.save(path)
    print(f"Exported to {path}")
    return path


if __name__ == "__main__":
    ym = sys.argv[1] if len(sys.argv) >= 2 else None
    results = screen(ym)
    if results:
        if not ym:
            with get_cursor(commit=False) as cur:
                cur.execute("SELECT MAX(year_month) AS ym FROM tw.monthly_revenue")
                ym = cur.fetchone()["ym"]
        export_excel(results, ym)
    else:
        print("No off-season outperformers found.")
