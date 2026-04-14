"""
Revenue Three Arrows stock screener (營收三支箭選股法).

Three conditions (all must be met):
  1. Short-term expansion: monthly revenue > historical average
  2. Momentum acceleration: monthly YoY% > cumulative YoY% (growth accelerating)
  3. Peak strength: monthly revenue is an all-time high

Usage:
  python -m analysis.revenue_three_arrows                # latest month
  python -m analysis.revenue_three_arrows 2026-03        # specific month
"""

import sys
from datetime import date
from decimal import Decimal

from db.connection import get_cursor


def screen(year_month: str = None) -> list[dict]:
    """
    Run the three-arrows screen for a given year_month (e.g. '2026-03').
    Returns list of qualifying stocks with details.
    """
    with get_cursor(commit=False) as cur:
        # Determine target month
        if not year_month:
            cur.execute("SELECT MAX(year_month) AS ym FROM tw.monthly_revenue")
            year_month = cur.fetchone()["ym"]

        print(f"Screening for {year_month} ...")

        year = int(year_month[:4])
        month = int(year_month[5:7])

        # Last year same month
        ly_year = year - 1
        ly_month_str = f"{ly_year}-{month:02d}"

        # Compute cumulative YoY: sum(revenue this year so far) vs sum(revenue last year same months)
        # Months this year: from {year}-01 to {year_month}
        ytd_start = f"{year}-01"
        ly_ytd_start = f"{ly_year}-01"
        ly_ytd_end = f"{ly_year}-{month:02d}"

        # Get all stocks with revenue in target month
        cur.execute("""
            SELECT r.stock_id, r.revenue, r.yoy_pct, r.mom_pct, r.note,
                   s.name, s.market, s.industry
            FROM tw.monthly_revenue r
            JOIN tw.stocks s ON r.stock_id = s.stock_id
            WHERE r.year_month = %s
              AND s.is_active = TRUE
        """, (year_month,))
        candidates = cur.fetchall()

        results = []

        for c in candidates:
            stock_id = c["stock_id"]
            revenue = c["revenue"]
            yoy_pct = float(c["yoy_pct"]) if c["yoy_pct"] is not None else None

            # --- Arrow 1: Revenue > historical average ---
            cur.execute("""
                SELECT AVG(revenue) AS avg_rev, COUNT(*) AS months
                FROM tw.monthly_revenue
                WHERE stock_id = %s AND year_month < %s
            """, (stock_id, year_month))
            hist = cur.fetchone()
            avg_rev = float(hist["avg_rev"]) if hist["avg_rev"] else None
            hist_months = hist["months"]

            if avg_rev is None or hist_months < 12:
                continue  # Not enough history

            arrow1 = revenue > avg_rev

            # --- Arrow 2: Monthly YoY > Cumulative YoY ---
            if yoy_pct is None:
                continue

            cur.execute("""
                SELECT SUM(revenue) AS ytd_sum
                FROM tw.monthly_revenue
                WHERE stock_id = %s AND year_month >= %s AND year_month <= %s
            """, (stock_id, ytd_start, year_month))
            ytd_sum = cur.fetchone()["ytd_sum"]

            cur.execute("""
                SELECT SUM(revenue) AS ly_ytd_sum
                FROM tw.monthly_revenue
                WHERE stock_id = %s AND year_month >= %s AND year_month <= %s
            """, (stock_id, ly_ytd_start, ly_ytd_end))
            ly_ytd_sum = cur.fetchone()["ly_ytd_sum"]

            if not ytd_sum or not ly_ytd_sum or ly_ytd_sum == 0:
                continue

            cum_yoy = round((ytd_sum - ly_ytd_sum) / ly_ytd_sum * 100, 2)
            arrow2 = yoy_pct > cum_yoy

            # --- Arrow 3: All-time high revenue ---
            cur.execute("""
                SELECT MAX(revenue) AS max_rev
                FROM tw.monthly_revenue
                WHERE stock_id = %s AND year_month < %s
            """, (stock_id, year_month))
            max_rev = cur.fetchone()["max_rev"] or 0
            arrow3 = revenue > max_rev

            # All three arrows
            if arrow1 and arrow2 and arrow3:
                results.append({
                    "stock_id": stock_id,
                    "name": c["name"],
                    "market": c["market"],
                    "industry": c["industry"],
                    "revenue": revenue,
                    "avg_revenue": round(avg_rev),
                    "rev_vs_avg_pct": round((revenue - avg_rev) / avg_rev * 100, 2),
                    "yoy_pct": yoy_pct,
                    "cum_yoy_pct": cum_yoy,
                    "prev_max_revenue": max_rev,
                    "mom_pct": float(c["mom_pct"]) if c["mom_pct"] is not None else None,
                    "note": c["note"],
                })

        # Sort by YoY descending
        results.sort(key=lambda x: x["yoy_pct"], reverse=True)
        print(f"Found {len(results)} stocks matching all three arrows.")
        return results


def export_excel(results: list[dict], year_month: str, path: str = None):
    """Export screening results to Excel."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, numbers

    if not path:
        path = f"revenue_three_arrows_{year_month}.xlsx"

    wb = Workbook()
    ws = wb.active
    ws.title = "Three Arrows"

    # Header style
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    headers = [
        ("代號", 10),
        ("名稱", 14),
        ("市場", 8),
        ("產業", 14),
        ("當月營收(千)", 16),
        ("歷史平均(千)", 16),
        ("超越均值%", 12),
        ("單月YoY%", 12),
        ("累計YoY%", 12),
        ("前歷史高(千)", 16),
        ("MoM%", 10),
        ("備註", 30),
    ]

    for col, (title, width) in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=title)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        ws.column_dimensions[cell.column_letter].width = width

    # Data rows
    num_fmt_int = '#,##0'
    num_fmt_pct = '#,##0.00'

    for row_idx, r in enumerate(results, 2):
        ws.cell(row=row_idx, column=1, value=r["stock_id"])
        ws.cell(row=row_idx, column=2, value=r["name"])
        ws.cell(row=row_idx, column=3, value=r["market"])
        ws.cell(row=row_idx, column=4, value=r["industry"])

        for col, key in [(5, "revenue"), (6, "avg_revenue"), (10, "prev_max_revenue")]:
            cell = ws.cell(row=row_idx, column=col, value=r[key])
            cell.number_format = num_fmt_int

        for col, key in [(7, "rev_vs_avg_pct"), (8, "yoy_pct"), (9, "cum_yoy_pct"), (11, "mom_pct")]:
            cell = ws.cell(row=row_idx, column=col, value=r[key])
            cell.number_format = num_fmt_pct

        ws.cell(row=row_idx, column=12, value=r["note"] or "")

    # Freeze header row
    ws.freeze_panes = "A2"

    # Auto-filter
    ws.auto_filter.ref = ws.dimensions

    wb.save(path)
    print(f"Exported to {path}")
    return path


if __name__ == "__main__":
    ym = sys.argv[1] if len(sys.argv) >= 2 else None
    results = screen(ym)

    if results:
        # Determine year_month for filename
        if not ym:
            with get_cursor(commit=False) as cur:
                cur.execute("SELECT MAX(year_month) AS ym FROM tw.monthly_revenue")
                ym = cur.fetchone()["ym"]
        export_excel(results, ym)
    else:
        print("No stocks matched all three arrows.")
